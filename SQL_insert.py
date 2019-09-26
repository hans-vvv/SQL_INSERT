from collections import defaultdict
from tkinter import Tk
from tkinter import filedialog
import os
import openpyxl


def info_from_xls(filename):

    insert_info = defaultdict(list) # All data from sheet
    
    # Read table names and scheme name
    wb = openpyxl.load_workbook(filename)
    table_names = wb.sheetnames
    scheme_name = filename.split('-')[1]
    insert_info['scheme_name'] = scheme_name

    # Read all tables
    for table_name in table_names:

        sheet = wb[table_name]

        #Read rows
        for row in sheet.iter_rows():
            table_row = [str(cell.value) for cell in row]
            insert_info[table_name].append(table_row)

    return insert_info


def create_insert_script(insert_info):

    with open('Insert-script.sql', 'w') as f:

        scheme_name = insert_info['scheme_name']
        print('use ' + scheme_name + ';', file=f)
             
        for table_name in insert_info:

            if table_name == 'scheme_name':
                continue

            col_header_count = len(insert_info[table_name][0])

            # cleanup column header names (Remove '-NQ')
            col_header = []
            for col_name in insert_info[table_name][0]:
                if '-NQ' in col_name:
                    col_name = col_name.split('-')[0]
                    col_header.append(col_name)
                else:
                   col_header.append(col_name) 

            # print SQL INSERT statements
            for index, row in enumerate(insert_info[table_name]):
                
                if index == 0:
                    continue

                # Format cells (Add quotes and NULL)
                formatted_cell_val = []
                print('INSERT INTO ' + table_name + ' (', end = '', file=f)
                for index, cell in enumerate(row):
                    if '-NQ' in insert_info[table_name][0][index]:
                        formatted_cell_val.append(cell)
                    elif cell is None:
                        formatted_cell_val.append('NULL')
                    else:   
                        formatted_cell_val.append("'" + cell + "'")

                # Print column headers 
                for index, cell in enumerate(col_header):
                    if index != col_header_count-1:
                        print(cell + ', ', end = '', file=f)
                    else:
                        print(cell, end = '', file=f)
                print(') VALUES (', end='', file=f)

                # Print formatted cells
                for index, cell in enumerate(formatted_cell_val):
                    if index != col_header_count-1:
                        print(cell + ', ', end = '', file=f)
                    else:
                        print(cell + ');', file=f)
              

def main():

    root = Tk()
    root.withdraw()
    filepath = filedialog.askopenfilename()
    directory, filename= os.path.split(filepath)
    os.chdir(directory)

    insert_info = info_from_xls(filename)
    
    create_insert_script(insert_info)


if __name__ == "__main__":
    main()

    






